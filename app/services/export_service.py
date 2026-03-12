"""Excel export service for weekly shift plans."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.models.shift import get_all_shifts

DAY_NAMES = ['Po', 'Út', 'St', 'Čt', 'Pá', 'So', 'Ne']

# Absence type colors (bg_hex, font_hex) - matching the web UI
ABSENCE_COLORS = {
    'dovolena':  ('FEF3C7', '92400E'),
    'nemoc':     ('FECACA', '991B1B'),
    'lekar':     ('E0E7FF', '3730A3'),
    'osobni':    ('D1FAE5', '065F46'),
    'nahradni':  ('F3E8FF', '6B21A8'),
    'jine':      ('E5E7EB', '374151'),
}

ABSENCE_LABELS = {
    'dovolena': 'Dovolená',
    'nemoc': 'Nemoc',
    'lekar': 'Lékař',
    'osobni': 'Osobní volno',
    'nahradni': 'Náhradní volno',
    'jine': 'Jiné',
}

# Shared style constants
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
HEADER_FILL = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
WEEKEND_FILL = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
SUMMARY_HEADER_FILL = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')


def generate_week_excel(plan, grid, dates, summary, task_summary):
    """Generate Excel workbook for weekly shift plan.

    Args:
        plan: dict with week_number, year, week_start
        grid: list of {employee: {name, ...}, days: [{date, assignment, ...}]}
        dates: list of 7 date objects
        summary: dict date_str → [{name, color, min_staff, staff_count}]
        task_summary: dict date_str → [{name, min_staff, staff_count}]

    Returns:
        bytes: xlsx file content
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f'Týden {plan["week_number"]}'

    # --- Page setup (A3 landscape, fit to 1 page) ---
    ws.page_setup.paperSize = 8  # A3
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.4

    # --- Row 1: Title ---
    date_from = dates[0].strftime('%d.%m.')
    date_to = dates[6].strftime('%d.%m.%Y')
    title = f'Plán směn {date_from}\u2013{date_to}  (týden {plan["week_number"]}/{plan["year"]})'

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True, color='1F2937')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # --- Row 2: blank spacer ---
    ws.row_dimensions[2].height = 6

    # --- Row 3: Header row ---
    row = 3
    _write_header(ws, row, dates)

    # --- Employee rows ---
    row = 4
    for grid_row in grid:
        _write_employee_row(ws, row, grid_row, dates)
        row += 1

    # --- Separator row ---
    sep_row = row
    ws.row_dimensions[sep_row].height = 6
    for col in range(1, 9):
        c = ws.cell(row=sep_row, column=col)
        c.fill = PatternFill(start_color='D1D5DB', end_color='D1D5DB', fill_type='solid')
    row += 1

    # --- Summary rows ---
    last_summary_row = _write_summary(ws, row, dates, summary, task_summary)

    # --- Shift legend ---
    _write_shift_legend(ws, last_summary_row + 1)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 20  # Employee name
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # --- Print header ---
    ws.oddHeader.center.text = f'Plán směn – týden {plan["week_number"]}/{plan["year"]}'
    ws.oddHeader.center.size = 8

    # Return as bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _write_header(ws, row, dates):
    """Write the header row: Employee + 7 day columns."""
    ws.row_dimensions[row].height = 32

    # Employee column header
    cell = ws.cell(row=row, column=1, value='Zaměstnanec')
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER

    # Day columns
    for i, d in enumerate(dates):
        day_label = f'{DAY_NAMES[i]} {d.strftime("%d.%m.")}'
        cell = ws.cell(row=row, column=i + 2, value=day_label)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Weekend header slightly different shade
        if d.weekday() >= 5:
            cell.fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
        else:
            cell.fill = HEADER_FILL


def _safe_get(obj, key, default=''):
    """Safely get a value from dict or sqlite3.Row."""
    try:
        val = obj[key]
        return val if val is not None else default
    except (KeyError, IndexError, TypeError):
        return default


def _write_employee_row(ws, row, grid_row, dates):
    """Write one employee row with 7 day cells."""
    ws.row_dimensions[row].height = 42

    # Employee name
    name_cell = ws.cell(row=row, column=1, value=grid_row['employee']['name'])
    name_cell.font = Font(bold=True, size=9, color='1F2937')
    name_cell.alignment = Alignment(vertical='center')
    name_cell.border = THIN_BORDER

    # Day cells
    for i, day in enumerate(grid_row['days']):
        col = i + 2
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        a = day.get('assignment') if day else None

        if a and _safe_get(a, 'is_absence'):
            _style_absence_cell(cell, a)
        elif a and (_safe_get(a, 'dept_name') or _safe_get(a, 'task_name') or _safe_get(a, 'shift_name')):
            _style_assignment_cell(cell, a)
        else:
            # Empty cell - weekend gets light gray
            if day.get('is_weekend'):
                cell.fill = WEEKEND_FILL
            cell.font = Font(size=8, color='CCCCCC')


def _style_absence_cell(cell, assignment):
    """Style a cell for absence display."""
    absence_type = _safe_get(assignment, 'absence_type', 'jine')
    label = ABSENCE_LABELS.get(absence_type, 'Jiné')
    bg_hex, font_hex = ABSENCE_COLORS.get(absence_type, ('E5E7EB', '374151'))

    parts = [label]
    note = _safe_get(assignment, 'note', '')
    if note:
        parts.append(note)

    # Partial absence (lékař with work assignment)
    dept_name = _safe_get(assignment, 'dept_name', '')
    task_name = _safe_get(assignment, 'task_name', '')
    if dept_name or task_name:
        work_parts = []
        if dept_name:
            work_parts.append(dept_name)
        if task_name:
            work_parts.append(task_name)
        parts.append(' · '.join(work_parts))

    cell.value = '\n'.join(parts)
    cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
    cell.font = Font(bold=True, size=9, color=font_hex)


def _style_assignment_cell(cell, a):
    """Style a cell for shift/department/task assignment."""
    parts = []

    dept_name = _safe_get(a, 'dept_name', '')
    task_name = _safe_get(a, 'task_name', '')
    shift_name = _safe_get(a, 'shift_name', '')
    note = _safe_get(a, 'note', '')

    if dept_name:
        parts.append(dept_name)
    if task_name:
        parts.append(task_name)
    if shift_name:
        parts.append(shift_name)
    if note:
        parts.append(f'({note})')

    cell.value = '\n'.join(parts)
    cell.font = Font(size=8, color='1F2937')

    # Department color background (light tint)
    dept_color = _safe_get(a, 'dept_color', '')
    if dept_color:
        # dept_color from DB is hex without #
        try:
            cell.fill = PatternFill(
                start_color=dept_color,
                end_color=dept_color,
                fill_type='solid'
            )
            # Use dark text on light backgrounds
            cell.font = Font(size=8, color='1F2937', bold=True)
        except (ValueError, TypeError):
            pass


def _write_summary(ws, row, dates, summary, task_summary):
    """Write staffing summary rows (department + task level). Returns last used row."""
    # Header for summary section
    ws.row_dimensions[row].height = 20
    header_cell = ws.cell(row=row, column=1, value='OBSAZENÍ')
    header_cell.font = Font(bold=True, size=9, color='374151')
    header_cell.fill = SUMMARY_HEADER_FILL
    header_cell.alignment = Alignment(vertical='center')
    header_cell.border = THIN_BORDER

    for col in range(2, 9):
        c = ws.cell(row=row, column=col)
        c.fill = SUMMARY_HEADER_FILL
        c.border = THIN_BORDER

    row += 1

    # Department-level summary
    # Get all department names from first day (they're the same for all days)
    first_day_key = dates[0].isoformat()
    dept_entries = summary.get(first_day_key, [])

    for dept in dept_entries:
        ws.row_dimensions[row].height = 20
        dept_label_cell = ws.cell(row=row, column=1, value=dept['name'])
        dept_label_cell.font = Font(bold=True, size=8, color='374151')
        dept_label_cell.alignment = Alignment(vertical='center')
        dept_label_cell.border = THIN_BORDER

        # Color indicator for department
        dept_color = _safe_get(dept, 'color', '')
        if dept_color:
            try:
                dept_label_cell.fill = PatternFill(
                    start_color=dept_color,
                    end_color=dept_color,
                    fill_type='solid'
                )
            except (ValueError, TypeError):
                pass

        for i, d in enumerate(dates):
            col = i + 2
            ds = d.isoformat()
            day_depts = summary.get(ds, [])

            # Find this department's data for this day
            dept_data = None
            for dd in day_depts:
                if dd['id'] == dept['id']:
                    dept_data = dd
                    break

            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if dept_data:
                count = dept_data['staff_count']
                min_s = dept_data['min_staff']
                cell.value = f'{count}/{min_s}'

                if count < min_s:
                    cell.font = Font(bold=True, size=9, color='DC2626')  # Red
                else:
                    cell.font = Font(size=9, color='16A34A')  # Green

            if d.weekday() >= 5:
                cell.fill = WEEKEND_FILL

        row += 1

    # Task-level summary (only tasks with min_staff > 0)
    first_tasks = task_summary.get(first_day_key, [])
    if first_tasks:
        for task in first_tasks:
            ws.row_dimensions[row].height = 18
            task_label_cell = ws.cell(row=row, column=1, value=f'  ↳ {task["name"]}')
            task_label_cell.font = Font(size=7, color='6B7280', italic=True)
            task_label_cell.alignment = Alignment(vertical='center')
            task_label_cell.border = THIN_BORDER

            for i, d in enumerate(dates):
                col = i + 2
                ds = d.isoformat()
                day_tasks = task_summary.get(ds, [])

                task_data = None
                for tt in day_tasks:
                    if tt['id'] == task['id']:
                        task_data = tt
                        break

                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

                if task_data:
                    count = task_data['staff_count']
                    min_s = task_data['min_staff']
                    cell.value = f'{count}/{min_s}'

                    if count < min_s:
                        cell.font = Font(bold=True, size=8, color='DC2626')
                    else:
                        cell.font = Font(size=8, color='16A34A')

                if d.weekday() >= 5:
                    cell.fill = WEEKEND_FILL

            row += 1

    return row


def _write_shift_legend(ws, row):
    """Write shift time legend below the summary."""
    shifts = get_all_shifts()
    if not shifts:
        return

    row += 1  # blank spacer row
    ws.row_dimensions[row].height = 8
    row += 1

    legend_parts = []
    for s in shifts:
        start = s['start_time'][:5] if s['start_time'] else ''
        end = s['end_time'][:5] if s['end_time'] else ''
        legend_parts.append(f'{s["name"]} = {start}–{end}')

    legend_text = 'Směny:  ' + '    |    '.join(legend_parts)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=legend_text)
    cell.font = Font(size=8, color='6B7280', italic=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')
