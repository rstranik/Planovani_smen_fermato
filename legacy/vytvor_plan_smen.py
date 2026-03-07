#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Plan smen - CSV z Airtable -> Formatovany Excel (A3 tisk)

Pouziti:
    1. Stahni CSV z Airtable
    2. Uloz ho do stejne slozky jako tento skript (nebo zadej cestu)
    3. Spust: python vytvor_plan_smen.py
    4. Otevri vysledny Excel a tiskni na A3

Autor: Automatizace pro FerMato
"""

import pandas as pd
import os
import sys
import glob
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# =============================================================================
# KONFIGURACE - UPRAV PODLE POTREBY
# =============================================================================

# Nazvy sloupcu v CSV (jak je exportuje Airtable)
# Pokud se u tebe jmenuji jinak, zmen to tady:
SLOUPEC_ZAMESTNANEC = "Zaměstnanec"
SLOUPEC_CAS_OD = "Čas od"
SLOUPEC_CAS_DO = "Čas do"
SLOUPEC_USEK = "Úsek"
SLOUPEC_PRACE = "Práce"
SLOUPEC_POZNAMKA = "Poznámka"

# Barevne schema podle useku (RGB hex bez #)
BARVY_USEKU = {
    "lahvování": "FFF2CC",   # svetle zluta
    "příchutě": "D9EAD3",   # svetle zelena
    "koření": "FCE5CD",     # svetle oranzova
    "expedice": "CFE2F3",   # svetle modra
    "sklad": "D9D9D9",      # svetle seda
    "závoz": "D9D9D9",      # svetle seda
    "rajčata": "F4CCCC",    # jemne cervena
    "sušičky": "EAD1DC",    # svetle ruzova
    "VINACZ": "D0E0E3",     # svetle tyrkysova
    "VEDS": "D9EAD3",       # svetle zelena
    "výroba": "FFF2CC",     # svetle zluta
    "VÝR": "FFF2CC",        # svetle zluta
    "EXP": "CFE2F3",        # svetle modra
}

# Hlavickove barvy
BARVA_HLAVICKA = "2E75B6"       # tmave modra
BARVA_HLAVICKA_TEXT = "FFFFFF"   # bila
BARVA_TITULEK = "1F4E79"        # velmi tmave modra

# Velikost pisma
FONT_TITULEK = 16
FONT_HLAVICKA = 11
FONT_CAS = 10
FONT_CINNOST = 9
FONT_POZNAMKA = 8
FONT_SOUCTY = 9


# =============================================================================
# HLAVNI LOGIKA
# =============================================================================

def najdi_csv_soubor():
    """Najde CSV soubor v aktualni slozce nebo se zepta uzivatele."""
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Hledame CSV soubory ve slozce skriptu
    csv_soubory = glob.glob(os.path.join(script_dir, "*.csv"))

    if not csv_soubory:
        print("CHYBA: Nebyl nalezen zadny CSV soubor ve slozce:")
        print(f"  {script_dir}")
        print("\nStahni CSV z Airtable a uloz ho do teto slozky.")
        sys.exit(1)

    if len(csv_soubory) == 1:
        print(f"Nalezen CSV: {os.path.basename(csv_soubory[0])}")
        return csv_soubory[0]

    # Vice CSV souboru - vyber nejnovejsi
    nejnovejsi = max(csv_soubory, key=os.path.getmtime)
    print(f"Nalezeno {len(csv_soubory)} CSV souboru, pouzivam nejnovejsi:")
    print(f"  {os.path.basename(nejnovejsi)}")
    return nejnovejsi


def nacti_data(csv_cesta):
    """Nacte CSV a pripravi data."""
    # Zkusime ruzne oddelovace
    for sep in [',', ';', '\t']:
        try:
            df = pd.read_csv(csv_cesta, sep=sep, encoding='utf-8')
            if len(df.columns) > 2:
                break
        except Exception:
            continue
    else:
        # Zkusime jeste s jinym kodovanim
        for sep in [',', ';', '\t']:
            try:
                df = pd.read_csv(csv_cesta, sep=sep, encoding='latin-1')
                if len(df.columns) > 2:
                    break
            except Exception:
                continue
        else:
            print("CHYBA: Nepodarilo se nacist CSV soubor.")
            print("Zkontroluj format souboru.")
            sys.exit(1)

    print(f"Nacteno {len(df)} zaznamu")
    print(f"Sloupce: {list(df.columns)}")

    # Overeni potrebnych sloupcu
    pozadovane = [SLOUPEC_ZAMESTNANEC, SLOUPEC_CAS_OD, SLOUPEC_CAS_DO,
                  SLOUPEC_USEK, SLOUPEC_PRACE]
    chybejici = [s for s in pozadovane if s not in df.columns]

    if chybejici:
        print(f"\nCHYBA: V CSV chybi sloupce: {chybejici}")
        print(f"Dostupne sloupce: {list(df.columns)}")
        print("\nUprav nazvy sloupcu v sekci KONFIGURACE na zacatku skriptu.")
        sys.exit(1)

    # Parsovani casu
    df[SLOUPEC_CAS_OD] = pd.to_datetime(df[SLOUPEC_CAS_OD], dayfirst=True)
    df[SLOUPEC_CAS_DO] = pd.to_datetime(df[SLOUPEC_CAS_DO], dayfirst=True)

    # Datum (bez casu)
    df['datum'] = df[SLOUPEC_CAS_OD].dt.date

    # Cas jako text
    df['cas_text'] = (df[SLOUPEC_CAS_OD].dt.strftime('%H:%M') + '–' +
                      df[SLOUPEC_CAS_DO].dt.strftime('%H:%M'))

    # Cinnost text
    df['cinnost_text'] = df[SLOUPEC_USEK].fillna('') + ' | ' + df[SLOUPEC_PRACE].fillna('')

    # Poznamka
    if SLOUPEC_POZNAMKA in df.columns:
        df[SLOUPEC_POZNAMKA] = df[SLOUPEC_POZNAMKA].fillna('')
    else:
        df[SLOUPEC_POZNAMKA] = ''

    return df


def zjisti_tyden(df):
    """Zjisti rozsah tydne z dat."""
    datumy = sorted(df['datum'].unique())
    pondeli = min(datumy)
    patek = max(datumy)

    # Generujeme vsechny dny pondeli-patek
    dny = []
    den = pondeli
    while den <= patek:
        # Preskocime vikend
        if den.weekday() < 5:  # 0=Po, 4=Pa
            dny.append(den)
        den += timedelta(days=1)

    return dny


def zjisti_barvu_useku(usek_text):
    """Vrati barvu podle useku."""
    if not usek_text:
        return None
    usek_lower = usek_text.lower()
    for klic, barva in BARVY_USEKU.items():
        if klic.lower() in usek_lower:
            return barva
    return None


NAZVY_DNU = {
    0: "Pondělí", 1: "Úterý", 2: "Středa", 3: "Čtvrtek", 4: "Pátek",
    5: "Sobota", 6: "Neděle"
}


def vytvor_excel(df, dny, vystupni_soubor):
    """Vytvori formatovany Excel soubor."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plán směn"

    # Styly
    tenka_cara = Side(style='thin', color='999999')
    border_all = Border(
        left=tenka_cara, right=tenka_cara,
        top=tenka_cara, bottom=tenka_cara
    )

    hlavicka_font = Font(name='Calibri', size=FONT_HLAVICKA, bold=True,
                         color=BARVA_HLAVICKA_TEXT)
    hlavicka_fill = PatternFill(start_color=BARVA_HLAVICKA,
                                end_color=BARVA_HLAVICKA, fill_type='solid')
    hlavicka_align = Alignment(horizontal='center', vertical='center',
                               wrap_text=True)

    cas_font = Font(name='Calibri', size=FONT_CAS, bold=True)
    cinnost_font = Font(name='Calibri', size=FONT_CINNOST)
    poznamka_font = Font(name='Calibri', size=FONT_POZNAMKA, italic=True,
                         color='666666')

    # ==========================================================
    # RADEK 1: Titulek
    # ==========================================================
    datum_od = min(dny).strftime('%d.%m.')
    datum_do = max(dny).strftime('%d.%m.%Y')
    cislo_tydne = min(dny).isocalendar()[1]

    titulek = f"Plán směn {datum_od}–{datum_do} (týden {cislo_tydne})"

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=1 + len(dny))
    titulek_cell = ws.cell(row=1, column=1, value=titulek)
    titulek_cell.font = Font(name='Calibri', size=FONT_TITULEK, bold=True,
                             color=BARVA_TITULEK)
    titulek_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # ==========================================================
    # RADEK 2: Prazdny (mezera)
    # ==========================================================
    ws.row_dimensions[2].height = 8

    # ==========================================================
    # RADEK 3: Hlavicka (Zamestnanec + dny)
    # ==========================================================
    radek_hlavicka = 3

    # Hlavicka "Zaměstnanec"
    cell = ws.cell(row=radek_hlavicka, column=1, value="Zaměstnanec")
    cell.font = hlavicka_font
    cell.fill = hlavicka_fill
    cell.alignment = hlavicka_align
    cell.border = border_all

    # Hlavicky dnu
    for i, den in enumerate(dny):
        nazev_dne = NAZVY_DNU.get(den.weekday(), "")
        datum_text = den.strftime('%d.%m.')
        hodnota = f"{nazev_dne}\n{datum_text}"

        cell = ws.cell(row=radek_hlavicka, column=2 + i, value=hodnota)
        cell.font = hlavicka_font
        cell.fill = hlavicka_fill
        cell.alignment = hlavicka_align
        cell.border = border_all

    ws.row_dimensions[radek_hlavicka].height = 35

    # ==========================================================
    # DATOVE RADKY: Zamestnanci a smeny
    # ==========================================================
    zamestnanci = sorted(df[SLOUPEC_ZAMESTNANEC].unique())
    print(f"Pocet zamestnancu: {len(zamestnanci)}")

    radek = radek_hlavicka + 1  # zaciname na radku 4

    for zamestnanec in zamestnanci:
        # Jmeno
        cell = ws.cell(row=radek, column=1, value=zamestnanec)
        cell.font = Font(name='Calibri', size=FONT_CAS, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='top',
                                   wrap_text=True)
        cell.border = border_all

        # Smeny pro kazdy den
        for i, den in enumerate(dny):
            zaznamy = df[
                (df[SLOUPEC_ZAMESTNANEC] == zamestnanec) &
                (df['datum'] == den)
            ]

            if zaznamy.empty:
                cell = ws.cell(row=radek, column=2 + i, value="")
                cell.border = border_all
                continue

            # Vezmeme prvni zaznam (nebo muzeme spojit vice)
            texty = []
            usek_pro_barvu = None

            for _, z in zaznamy.iterrows():
                cas = z['cas_text']
                cinnost = z['cinnost_text']
                poznamka = z.get(SLOUPEC_POZNAMKA, '')

                text = f"{cas}\n{cinnost}"
                if poznamka and str(poznamka).strip():
                    text += f"\n{poznamka}"

                texty.append(text)

                if usek_pro_barvu is None:
                    usek_pro_barvu = str(z.get(SLOUPEC_USEK, ''))

            hodnota = "\n---\n".join(texty) if len(texty) > 1 else texty[0]

            cell = ws.cell(row=radek, column=2 + i, value=hodnota)
            cell.alignment = Alignment(horizontal='left', vertical='top',
                                       wrap_text=True)
            cell.border = border_all

            # Barva podle useku
            barva = zjisti_barvu_useku(usek_pro_barvu)
            if barva:
                cell.fill = PatternFill(start_color=barva, end_color=barva,
                                        fill_type='solid')
                cell.font = cinnost_font
            else:
                cell.font = cinnost_font

        ws.row_dimensions[radek].height = 52
        radek += 1

    # ==========================================================
    # SOUCTY - Pocet lidi podle useku na den
    # ==========================================================
    radek += 1  # prazdny radek

    # Nadpis souctu
    cell = ws.cell(row=radek, column=1, value="SOUČTY")
    cell.font = Font(name='Calibri', size=FONT_SOUCTY, bold=True,
                     color=BARVA_TITULEK)
    cell.border = border_all
    radek += 1

    # Zjistime vsechny useky
    useky = sorted(df[SLOUPEC_USEK].dropna().unique())

    # Celkovy pocet
    cell = ws.cell(row=radek, column=1, value="Celkem")
    cell.font = Font(name='Calibri', size=FONT_SOUCTY, bold=True)
    cell.border = border_all

    for i, den in enumerate(dny):
        pocet = len(df[df['datum'] == den][SLOUPEC_ZAMESTNANEC].unique())
        cell = ws.cell(row=radek, column=2 + i, value=pocet)
        cell.font = Font(name='Calibri', size=FONT_SOUCTY, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_all

    ws.row_dimensions[radek].height = 20
    radek += 1

    # Pocty podle useku
    for usek in useky:
        cell = ws.cell(row=radek, column=1, value=usek)
        cell.font = Font(name='Calibri', size=FONT_SOUCTY)
        cell.border = border_all

        barva = zjisti_barvu_useku(usek)
        if barva:
            cell.fill = PatternFill(start_color=barva, end_color=barva,
                                    fill_type='solid')

        for i, den in enumerate(dny):
            pocet = len(df[
                (df['datum'] == den) &
                (df[SLOUPEC_USEK] == usek)
            ][SLOUPEC_ZAMESTNANEC].unique())

            cell = ws.cell(row=radek, column=2 + i, value=pocet if pocet > 0 else "")
            cell.font = Font(name='Calibri', size=FONT_SOUCTY)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_all

            if barva:
                cell.fill = PatternFill(start_color=barva, end_color=barva,
                                        fill_type='solid')

        ws.row_dimensions[radek].height = 18
        radek += 1

    # ==========================================================
    # SLOUPCE - SIRKY
    # ==========================================================
    ws.column_dimensions['A'].width = 22  # Jmena

    for i in range(len(dny)):
        col_letter = get_column_letter(2 + i)
        ws.column_dimensions[col_letter].width = 22

    # ==========================================================
    # NASTAVENI TISKU - A3 NA SIRKU
    # ==========================================================
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = None

    # Okraje (v palcich) - uzke
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    # Opakovat hlavicku pri tisku
    ws.print_title_rows = f'{radek_hlavicka}:{radek_hlavicka}'

    # ==========================================================
    # ULOZENI
    # ==========================================================
    wb.save(vystupni_soubor)
    print(f"\n{'='*50}")
    print(f"HOTOVO! Soubor ulozen:")
    print(f"  {vystupni_soubor}")
    print(f"{'='*50}")
    print(f"\nOtevri soubor v Excelu a tiskni na A3 (Na sirku).")
    print(f"Nastaveni tisku je predpripravene.")


def main():
    print("=" * 50)
    print("  PLAN SMEN - CSV -> Excel (A3 tisk)")
    print("=" * 50)
    print()

    # Najdi CSV
    if len(sys.argv) > 1:
        csv_cesta = sys.argv[1]
    else:
        csv_cesta = najdi_csv_soubor()

    if not os.path.exists(csv_cesta):
        print(f"CHYBA: Soubor nenalezen: {csv_cesta}")
        sys.exit(1)

    # Nacti data
    print(f"\nNaciitam data z: {os.path.basename(csv_cesta)}")
    df = nacti_data(csv_cesta)

    # Zjisti tyden
    dny = zjisti_tyden(df)
    print(f"Rozsah: {dny[0].strftime('%d.%m.%Y')} - {dny[-1].strftime('%d.%m.%Y')}")
    print(f"Pocet dnu: {len(dny)}")

    # Vystupni soubor
    script_dir = os.path.dirname(os.path.abspath(__file__))
    datum_od = min(dny).strftime('%d.%m.')
    datum_do = max(dny).strftime('%d.%m.%Y')
    vystupni = os.path.join(
        script_dir,
        f"Plan_smen_{datum_od}-{datum_do}.xlsx"
    )

    # Vytvor Excel
    vytvor_excel(df, dny, vystupni)

    # Otevri soubor (Windows)
    try:
        os.startfile(vystupni)
    except Exception:
        pass


if __name__ == "__main__":
    main()
